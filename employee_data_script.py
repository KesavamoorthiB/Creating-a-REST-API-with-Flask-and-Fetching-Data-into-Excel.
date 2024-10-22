import openpyxl
import requests
import os

# Set the Excel file name
file_name = 'employee_data.xlsx'

# Check if the Excel file exists, if not, create it
if not os.path.exists(file_name):
    workbook = openpyxl.Workbook()
    workbook.save(file_name)

# Load the Excel file
workbook = openpyxl.load_workbook(file_name)
sheet = workbook.active

# Check if headers exist, otherwise, create them
headers = ['EmployeeID', 'Name', 'Department', 'Joining Date', 'Email']
if sheet.max_row == 1:  # If only one row (possibly headers)
    sheet.append(headers)
    workbook.save(file_name)

# REST API URL (replace with your actual API)
api_url = api_url = "http://127.0.0.1:5000/employees"
 

# Fetch employee data from the REST API
response = requests.get(api_url)
if response.status_code == 200:
    employee_data = response.json()  # Assuming the API returns JSON
    for employee in employee_data:
        row = [
            employee['id'],
            employee['name'],
            employee['department'],
            employee['joining_date'],
            employee['email']
        ]
        sheet.append(row)  # Append the new employee data

    # Save the Excel file after appending the data
    workbook.save(file_name)
    print("Employee data written to Excel successfully.")
else:
    print(f"Failed to fetch data. Status code: {response.status_code}")
